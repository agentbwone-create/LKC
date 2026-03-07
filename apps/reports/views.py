import io
from collections import defaultdict
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from apps.exams.models import ExamResult
from apps.core.models import Subject, Exam


def compute_grade(percentage):
    if percentage >= 90: return 'A*'
    if percentage >= 80: return 'A'
    if percentage >= 70: return 'B'
    if percentage >= 60: return 'C'
    if percentage >= 50: return 'D'
    if percentage >= 40: return 'E'
    return 'U'


class HODReportView(APIView):
    """
    Generate HOD Grades Analysis Excel report.
    Matches the exact template format from HOD_GRADES_ANALYSIS_TEMPLATE_2.xlsx
    POST /api/reports/hod-report/
    {
        "exam_id": 1,
        "subject_id": 2,
        "pass_percentage": 50,
        "credit_pass_percentage": 60,
        "report_title": "ICT Department Term 1 2026"
    }
    """

    def post(self, request):
        exam_id = request.data.get('exam_id')
        subject_id = request.data.get('subject_id')
        pass_pct = float(request.data.get('pass_percentage', 50))
        credit_pct = float(request.data.get('credit_pass_percentage', 60))
        report_title = request.data.get('report_title', '')

        if not exam_id or not subject_id:
            return Response({'error': 'exam_id and subject_id are required.'}, status=400)

        try:
            from apps.exams.models import Exam as ExamModel
            exam = ExamModel.objects.get(pk=exam_id)
            subject = Subject.objects.get(pk=subject_id)
        except (ExamModel.DoesNotExist, Subject.DoesNotExist):
            return Response({'error': 'Exam or Subject not found.'}, status=404)

        results = ExamResult.objects.filter(
            exam=exam,
            subject=subject,
            state__in=['confirmed', 'teacher_approved', 'done'],
        ).select_related('student', 'division', 'teacher', 'exam__form')

        if not results.exists():
            return Response({'error': 'No confirmed results found.'}, status=404)

        title = report_title or f'{subject.name} Department {exam.exam_name}'
        xlsx_data = self._build_xlsx(results, title, pass_pct, credit_pct, subject, exam)

        filename = f'{subject.name.replace(" ", "_")}_{exam.exam_name.replace(" ", "_")}_HOD_Report.xlsx'
        response = HttpResponse(
            xlsx_data,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _build_xlsx(self, results, title, pass_pct, credit_pct, subject, exam):
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'HOD Grades Analysis'

        GRADE_COLS = ['A*', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'U']

        # Styles
        title_font = Font(name='Arial', bold=True, size=16)
        subtitle_font = Font(name='Arial', size=14)
        header_font = Font(name='Arial', size=12)
        data_font = Font(name='Arial', bold=True, size=14)
        formula_font = Font(name='Arial', size=12)
        note_font = Font(name='Arial', size=11)
        center = Alignment(horizontal='center', vertical='center')
        thin = Side(style='thin')
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Column widths
        for col, w in [('A',3),('B',15),('C',10),('D',5),('E',5),('F',5),('G',5),
                        ('H',5),('I',5),('J',5),('K',5),('L',5),('M',7),
                        ('N',16),('O',10),('P',16),('Q',25)]:
            ws.column_dimensions[col].width = w

        # Row 1 - Title
        ws.row_dimensions[1].height = 21
        ws.merge_cells('B1:O1')
        ws['B1'].value = title
        ws['B1'].font = title_font
        ws['B1'].alignment = center
        ws['P1'].value = 'Need for both assessments- End of Term and Mid Term assessment'
        ws['P1'].font = note_font

        # Row 2 - Subtitle
        ws.row_dimensions[2].height = 18.6
        ws.merge_cells('B2:O2')
        ws['B2'].value = 'Teacher to Teacher Analysis'
        ws['B2'].font = subtitle_font
        ws['B2'].alignment = center
        ws['P2'].value = f'Pass%={int(pass_pct)}%'
        ws['P2'].font = note_font
        ws['Q2'].value = f'Credit pass = {int(credit_pct)}% and above'
        ws['Q2'].font = note_font

        # Group results: form -> (division, teacher) -> grade_counts
        data = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
        form_order = []
        seen_forms = set()

        for r in results:
            form_name = r.exam.form.name if r.exam.form else 'Unknown'
            div_display = r.division.name if r.division else 'Unknown'
            teacher_name = r.teacher.teacher_name if r.teacher else 'Unknown'
            g = compute_grade(r.percentage)

            if form_name not in seen_forms:
                form_order.append(form_name)
                seen_forms.add(form_name)
            data[form_name][(div_display, teacher_name)][g] += 1

        current_row = 3

        def hcell(row, col, val):
            c = ws.cell(row=row, column=col, value=val)
            c.font = header_font
            c.alignment = center
            c.border = thin_border

        for form_name in form_order:
            form_data = data[form_name]
            ws.row_dimensions[current_row].height = 15.6

            hcell(current_row, 2, 'Teacher')
            hcell(current_row, 3, form_name)
            for i, g in enumerate(GRADE_COLS):
                hcell(current_row, 4 + i, g)
            hcell(current_row, 13, 'Total')
            hcell(current_row, 14, 'Credit Pass %')
            hcell(current_row, 15, 'Pass %')
            current_row += 1
            data_start_row = current_row

            for (div_display, teacher_name), grade_counts in sorted(form_data.items()):
                ws.row_dimensions[current_row].height = 18
                r = current_row

                for col, val, fnt in [(2, teacher_name, header_font), (3, div_display, header_font)]:
                    c = ws.cell(row=r, column=col, value=val)
                    c.font = fnt; c.alignment = center; c.border = thin_border

                for i, g in enumerate(GRADE_COLS):
                    c = ws.cell(row=r, column=4+i, value=grade_counts.get(g, 0))
                    c.font = data_font; c.alignment = center; c.border = thin_border

                ws[f'M{r}'] = f'=SUM(D{r}:L{r})'
                ws[f'M{r}'].font = formula_font; ws[f'M{r}'].alignment = center; ws[f'M{r}'].border = thin_border

                ws[f'N{r}'] = f'=IF(M{r}=0,0,SUM(D{r}:G{r})/M{r})'
                ws[f'N{r}'].font = formula_font; ws[f'N{r}'].alignment = center
                ws[f'N{r}'].number_format = '0.0%'; ws[f'N{r}'].border = thin_border

                ws[f'O{r}'] = f'=IF(M{r}=0,0,SUM(D{r}:I{r})/M{r})'
                ws[f'O{r}'].font = formula_font; ws[f'O{r}'].alignment = center
                ws[f'O{r}'].number_format = '0.0%'; ws[f'O{r}'].border = thin_border

                current_row += 1

            data_end_row = current_row - 1
            r = current_row
            ws.row_dimensions[r].height = 15.6

            c = ws.cell(row=r, column=3, value='Total')
            c.font = formula_font; c.border = thin_border

            for i in range(9):
                cl = get_column_letter(4 + i)
                cell = ws.cell(row=r, column=4+i)
                cell.value = f'=SUM({cl}{data_start_row}:{cl}{data_end_row})'
                cell.font = formula_font; cell.alignment = center; cell.border = thin_border

            ws[f'M{r}'] = f'=SUM(M{data_start_row}:M{data_end_row})'
            ws[f'M{r}'].font = formula_font; ws[f'M{r}'].alignment = center; ws[f'M{r}'].border = thin_border

            ws[f'N{r}'] = f'=IF(M{r}=0,0,SUM(D{r}:G{r})/M{r})'
            ws[f'N{r}'].font = formula_font; ws[f'N{r}'].alignment = center
            ws[f'N{r}'].number_format = '0.0%'; ws[f'N{r}'].border = thin_border

            ws[f'O{r}'] = f'=IF(M{r}=0,0,SUM(D{r}:I{r})/M{r})'
            ws[f'O{r}'].font = formula_font; ws[f'O{r}'].alignment = center
            ws[f'O{r}'].number_format = '0.0%'; ws[f'O{r}'].border = thin_border

            current_row += 2  # spacer

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()


class ReportPreviewView(APIView):
    """Return HOD report data as JSON for frontend preview before download."""

    def post(self, request):
        exam_id = request.data.get('exam_id')
        subject_id = request.data.get('subject_id')

        if not exam_id or not subject_id:
            return Response({'error': 'exam_id and subject_id are required.'}, status=400)

        results = ExamResult.objects.filter(
            exam_id=exam_id,
            subject_id=subject_id,
            state__in=['confirmed', 'teacher_approved', 'done'],
        ).select_related('student', 'division', 'teacher', 'exam__form')

        GRADE_COLS = ['A*', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'U']
        data = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
        form_order = []
        seen_forms = set()

        for r in results:
            form_name = r.exam.form.name if r.exam.form else 'Unknown'
            div = r.division.name if r.division else 'Unknown'
            teacher = r.teacher.teacher_name if r.teacher else 'Unknown'
            g = compute_grade(r.percentage)
            if form_name not in seen_forms:
                form_order.append(form_name)
                seen_forms.add(form_name)
            data[form_name][(div, teacher)][g] += 1

        output = []
        for form_name in form_order:
            rows = []
            for (div, teacher), grades in sorted(data[form_name].items()):
                total = sum(grades.get(g, 0) for g in GRADE_COLS)
                credit = round(sum(grades.get(g, 0) for g in ['A*','A','B','C']) / total * 100, 1) if total else 0
                pass_ = round(sum(grades.get(g, 0) for g in ['A*','A','B','C','D','E']) / total * 100, 1) if total else 0
                rows.append({
                    'teacher': teacher, 'division': div,
                    'grades': {g: grades.get(g, 0) for g in GRADE_COLS},
                    'total': total, 'credit_pct': credit, 'pass_pct': pass_,
                })
            output.append({'form': form_name, 'rows': rows})

        return Response({'forms': output, 'total_results': results.count()})


class ClassTeacherPerformanceView(APIView):
    """
    Returns overall student performance across all subjects for a class teacher's division.

    The class teacher is identified by their FacultyInfo linked to the logged-in user,
    OR by passing ?teacher_id=<id> (for admin use).
    Also accepts ?division_id=<id> and ?exam_id=<id> as optional filters.

    GET /api/reports/class-performance/
    GET /api/reports/class-performance/?division_id=2&exam_id=3
    GET /api/reports/class-performance/?teacher_id=5  (admin use)

    Response:
    {
        "division": "Form 3 A",
        "exam": "Term 1 2025",   // or "All Exams" if no filter
        "students": [
            {
                "student_id": 1,
                "student_name": "Jane Doe",
                "enrollment_number": "S001",
                "subjects": [
                    {
                        "subject": "Mathematics",
                        "obtain_marks": 78,
                        "maximum_marks": 100,
                        "percentage": 78.0,
                        "grade": "B",
                        "result": "pass"
                    },
                    ...
                ],
                "overall_average": 72.5,
                "overall_grade": "B",
                "subjects_passed": 5,
                "subjects_failed": 1,
                "total_subjects": 6
            },
            ...
        ]
    }
    """

    def get(self, request):
        from apps.faculty.models import FacultyInfo
        from apps.core.models import SchoolDivision
        from apps.students.models import StudentInfo

        # Resolve the teacher
        teacher_id = request.query_params.get('teacher_id')
        division_id = request.query_params.get('division_id')
        exam_id = request.query_params.get('exam_id')

        if teacher_id:
            try:
                teacher = FacultyInfo.objects.get(pk=teacher_id)
            except FacultyInfo.DoesNotExist:
                return Response({'error': 'Teacher not found.'}, status=404)
        elif request.user and request.user.is_authenticated:
            teacher = getattr(request.user, 'faculty_profile', None)
            if not teacher:
                return Response({'error': 'No faculty profile linked to this user.'}, status=403)
        else:
            return Response({'error': 'teacher_id is required or user must be authenticated.'}, status=400)

        # Resolve the division
        if division_id:
            try:
                division = SchoolDivision.objects.get(pk=division_id)
            except SchoolDivision.DoesNotExist:
                return Response({'error': 'Division not found.'}, status=404)
        else:
            # Use the teacher's assigned class teacher division
            division = SchoolDivision.objects.filter(teacher=teacher).first()
            if not division:
                return Response({'error': 'This teacher is not assigned as a class teacher for any division.'}, status=404)

        # Get all students in this division
        students = StudentInfo.objects.filter(division=division, state='active').order_by('student_full_name')

        if not students.exists():
            return Response({
                'division': str(division),
                'exam': 'N/A',
                'students': [],
                'message': 'No active students in this division.'
            })

        # Build results query
        results_qs = ExamResult.objects.filter(
            student__in=students,
        ).select_related('student', 'subject', 'exam')

        if exam_id:
            from apps.exams.models import Exam as ExamModel
            try:
                exam_obj = ExamModel.objects.get(pk=exam_id)
            except ExamModel.DoesNotExist:
                return Response({'error': 'Exam not found.'}, status=404)
            results_qs = results_qs.filter(exam=exam_obj)
            exam_label = exam_obj.exam_name
        else:
            exam_label = 'All Exams'

        # Group results by student
        from collections import defaultdict
        student_results = defaultdict(list)
        for result in results_qs:
            student_results[result.student_id].append(result)

        output_students = []
        for student in students:
            s_results = student_results.get(student.id, [])

            subjects_data = []
            for r in s_results:
                subjects_data.append({
                    'subject_id': r.subject_id,
                    'subject': r.subject.name,
                    'exam': r.exam.exam_name,
                    'obtain_marks': r.obtain_marks,
                    'maximum_marks': r.maximum_marks,
                    'percentage': r.percentage,
                    'grade': r.grade,
                    'result': r.result,
                })

            if subjects_data:
                avg = round(sum(s['percentage'] for s in subjects_data) / len(subjects_data), 2)
                overall_grade = compute_grade(avg)
                passed = sum(1 for s in subjects_data if s['result'] == 'pass')
                failed = len(subjects_data) - passed
            else:
                avg = 0.0
                overall_grade = 'N/A'
                passed = 0
                failed = 0

            output_students.append({
                'student_id': student.id,
                'student_name': student.student_full_name,
                'enrollment_number': student.enrollment_number,
                'gender': student.gender,
                'subjects': subjects_data,
                'overall_average': avg,
                'overall_grade': overall_grade,
                'subjects_passed': passed,
                'subjects_failed': failed,
                'total_subjects': len(subjects_data),
            })

        # Sort by overall average descending (ranking)
        output_students.sort(key=lambda x: x['overall_average'], reverse=True)
        for i, s in enumerate(output_students, 1):
            s['rank'] = i

        return Response({
            'division': str(division),
            'division_id': division.id,
            'exam': exam_label,
            'total_students': len(output_students),
            'students': output_students,
        })
